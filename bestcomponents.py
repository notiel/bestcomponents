import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials
import sys

import requests
from bs4 import BeautifulSoup
from bs4 import Tag

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl import load_workbook

@dataclass
class Product:
    id: str
    actual: int
    delivery: int
    prognosis: int
    prognosis_type: str
    prices_actual: dict
    prices_delivery: dict
    partnumber: str

terra_base = r"https://www.terraelectronica.ru/"
onelec_base = r'https://onelec.ru/products/'
BIG_PRICE = 10000
maxline = 100

CREDENTIALS_FILE = 'LSComponents.json'


def get_search_links_for_position(cathegory: str, value: str, footprint: str) -> [str]:
    """
    :param cathegory: position type
    :param value:  position value
    :param footprint:  position footprint
    :return: searchlinks for this position
    """
    search_links = []
    if cathegory == 'RES':
        position = value + ' ' + footprint
        position += ' 1%'
        search_links = get_search_links_from_page(position)
    if cathegory == 'CAP':
        position = value + ' ' + footprint
        if 'u' or 'n' in value:
            search_links = get_search_links_from_page(position + ' x7r')
            search_links.extend(get_search_links_from_page(position + ' x5r'))
    for link in search_links:
        if '0603' in position:
            link = correct_link_for_0603(link)
    return search_links


def get_search_links_from_page(search_text) -> [str]:
    """
    function gets list of search links for selected query with "smd" in description
    :param search_text - query to search (10u 16V 0805) for example
    :return: list of links with searc results
    """
    search_query = "+".join(search_text.split())
    url = terra_base + "search?text=" + search_query
    r = requests.get(url)
    soup = BeautifulSoup(r.text)
    links = soup.find('ul', {'class': "search-list"})
    try:
        search_links = [link.contents for link in links.contents if isinstance(link, Tag)]
    except AttributeError:
        raise
    real_search_links = []
    for link in search_links:
        for tag in link:
            if isinstance(tag, Tag):
                search_string = tag.contents[0]
                if 'SMD' in search_string:
                    real_search_links.append(tag.attrs['href'])
    return real_search_links


def correct_link_for_0603(link: str) -> str:
    """
    0603 cage has metric and nonmetric varieties. This function excludes metric cage
    :param link: search link
    :return: corrected link withoot metric 0603 (aka 0201)
    """
    query = link.split('%26')
    query = [q for q in query if '0201' not in q]
    link = '%26'.join(query)
    return link


def get_product_list(link: str) -> [str]:
    """
    function gets products ids using search link
    :param link: search link
    :return: list of product ids
    """
    url = terra_base + link
    r = requests.get(url)
    soup = BeautifulSoup(r.text)
    pages = soup.findAll('li', {'class': 'waves-effect'})
    products = []
    if pages:
        for page in set(pages):
            url = terra_base + page.contents[0].attrs['href']
            r = requests.get(url)
            soup = BeautifulSoup(r.text)
            links = soup.findAll('td', {'class': 'table-item-name'})
            products.extend([link.attrs['data-code'] for link in links])
        return products
    links = soup.findAll('td', {'class': 'table-item-name'})
    products = [link.attrs['data-code'] for link in links]
    return products


def get_actual_info(product_id: str) -> (int, dict, str):
    """
    function gets actual price and quantity of product. If on demand only return 0 and {}
    :param product_id: product id
    :return: quantity, dictionary with prices, partnumber
    """
    url = terra_base + "product/" + product_id
    res = requests.get(url)
    soup = BeautifulSoup(res.text)
    actual = soup.find('div', {'class': 'box-title'})
    partnumber = soup.find('h1', {'class': 'truncate'})
    partnumber = partnumber.contents[0].split()[0]
    if actual:
        actual = [tag for tag in actual if isinstance(tag, Tag)]
        actual_quantity = int(actual[0].contents[0].replace("шт.", ""))
        price_data = [tag for tag in soup.find('span', {'class': 'prices'}) if isinstance(tag, Tag)]
        prices_actual = {}
        for price in price_data:
            prices_actual[int(price.attrs['data-count'])] = float(price.attrs['data-price'])
        return actual_quantity, prices_actual, partnumber
    return 0, {}, partnumber


def get_delivery_info(product_id: str) -> (int, int, str, dict):
    """
    function gets delivery data for product
    :param product_id: id of product
    :return: quantity available, number of delivery units, delivery unit: day or week, delivery prices
    """
    data = '{"jsonrpc":"2.0","method":"update_offers","params":{"code":%s},"id":"objUpdateOffers||1"}' % product_id
    response = requests.post('https://www.terraelectronica.ru/services', data=data)
    res = response.text
    # print(product_id)
    res = res.split('"best_offer":')[1]
    res = res.replace(r'\"', r'"')
    res = res.replace("\n", "")
    soup = BeautifulSoup(res)
    delivery_data = soup.find('div', {'class': 'box-title'})
    if delivery_data:
        delivery_data = [tag for tag in soup.find('div', {'class': 'box-title'}) if isinstance(tag, Tag)]
        actual = delivery_data[0]
        if 'ПОД ЗАКАЗ' in actual.contents[0]:
            quantity = actual.contents[1].contents[0]
            quantity = int(quantity.replace("шт.", ""))
            prognosis = delivery_data[1].contents[0]
            if "недел" in prognosis:
                prognosis_type = "Weeks"
                if 'более' not in prognosis:
                    prognosis = int(prognosis.split()[2].split('-')[0])
                else:
                    prognosis = 100
            else:
                if "дн" in prognosis:
                    prognosis_type = "Days"
                    prognosis = int(prognosis.split()[2])
        price_data = [tag for tag in soup.find('span', {'class': 'prices'}) if isinstance(tag, Tag)]
        prices_delivery = {}
        for price in price_data:
            prices_delivery[int(price.attrs['data-count'])] = float(price.attrs['data-price'])
        return quantity, prognosis, prognosis_type, prices_delivery
    return 0, 0, None, {}


def get_product_data(link: str, products: [Product]):
    """
    adds product data to product list
    :param link: link with product links
    :param products: list of products already got
    :return:
    """
    product_ids = get_product_list(link)
    for product_id in product_ids:
        actual, prices_actual, partnumber = get_actual_info(product_id)
        delivery, prognosis, prognosis_type, prices_delivery = get_delivery_info(product_id)
        products.append(
            Product(id=product_id, actual=actual, delivery=delivery, prices_actual=prices_actual,
                    prices_delivery=prices_delivery, prognosis=prognosis,
                    prognosis_type=prognosis_type, partnumber=partnumber))
    return


def get_min_price_actual_with_quantity(products: [Product], quantity: int) -> (str, float):
    """
    gets actual offer for position witn minimal price and not less then quantity items (price must be chosen
    for required quantity)
    :param quantity: required quantity of items
    :param products: list with offers for this position
    :return: id of best offer, price of best offer
    """
    actual_prices = {}
    for product in products:
        if product.actual >= quantity:
            min_price = product.prices_actual[1]
            min_id = product.id
            for q in product.prices_actual.keys():
                if q <= quantity and min_price >= product.prices_actual[q]:
                    min_price = product.prices_actual[q]
            actual_prices[product.id] = min_price
    if actual_prices:
        for product in actual_prices.keys():
            if actual_prices[product] < min_price:
                min_id = product
                min_price = actual_prices[product]
        return min_id, min_price
    return "0", -1


def get_min_price_quantity_data(products: [Product], quantity: int, date: int) -> (float, str, int):
    """
    get best offer for quanity units with no more then date days of delivery
    :param products: list if offers
    :param date: max days of delivery
    :param quantity: required qiantity
    :return: best price, id of best offer, days of delivery
    """
    min_id, min_price_actual = get_min_price_actual_with_quantity(products, quantity)
    delivery_prices = {}
    for product in products:
        if product.delivery >= quantity:
            prognosis = product.prognosis if product.prognosis_type == "Days" else product.prognosis * 7
            if prognosis <= date:
                min_price = BIG_PRICE
                for q in product.prices_delivery.keys():
                    if q <= quantity and min_price >= product.prices_delivery[q]:
                        min_price = product.prices_delivery[q]
                        delivery_prices[product.id] = [min_price, prognosis]
    if delivery_prices:
        min_delivery_price = min_price
        for product in delivery_prices.keys():
            if delivery_prices[product][0] < min_delivery_price:
                min_delivery_id = product
                min_delivery_price = delivery_prices[product][0]
                min_delivery_prognosis = delivery_prices[product][1]
    else:
        return min_price_actual, min_id, 1
    if min_price == 0:
        return min_delivery_price, min_delivery_id, min_delivery_prognosis
    if min_price_actual <= min_delivery_price:
        return min_price_actual, min_id, 1
    else:
        return min_delivery_price, min_delivery_id, min_delivery_prognosis



def get_terra_by_pn(partnumber:str) -> (float, str):
    """
    gets data from terra by partnumber
    :param partnumber:
    :return: price, url
    """
    url = terra_base + "search?text=" + partnumber
    res = requests.get(url)
    terra_url = ""
    terra_price = 0
    if 'product' in res.url:
        terra_url = res.url
        soup = BeautifulSoup(res.text)
        tags = soup.find('div', {'class': 'fast-buy'})
        if tags:
            tag = soup.find('span', {'class': 'price-single price-active'})
            terra_price = float(tag.attrs['data-price'])
    return terra_price, terra_url


def get_onelec_pn(partnumber: str) -> (float, str):
    """
    gets url and price from onelec
    :param partnumber: partnumber of product
    :return: price, url
    """
    url = onelec_base + partnumber
    res = requests.get(url)
    onelec_url = ""
    onelec_price = 0
    if res.status_code != 404:
        onelec_url = url
        soup = BeautifulSoup(res.text)
        table = soup.find('table', {'class': "table product-offers"})
        try:
            for tag in [tag for tag in table.contents[0].contents if isinstance(tag, Tag)]:
                try:
                    delivery = int(tag.contents[0].text.split()[1])
                    if delivery <= 5 and 'по запросу' not in tag.contents[1].text:
                        price = float(
                            tag.contents[2].contents[0].contents[0]['data-price-rub'].split()[0].replace(',', '.'))
                        if onelec_price == 0:
                            onelec_price = price
                        else:
                           if price < onelec_price:
                               onelec_price = price
                except ValueError:
                    continue
        except AttributeError:
            pass
    return onelec_price, onelec_url


def get_best_price_from_onelec_terra_by_pn(partnumber: str)->(float, str, str):
    """
    function gets best price from onelec and terra by partnumber and selects best
    :param partnumber: partnumber for spreadsheet
    :return: best price, best url, other url+price
    """
    terra_price, terra_url = get_terra_by_pn(partnumber)
    onelec_price, onelec_url = get_onelec_pn(partnumber)
    if terra_price < onelec_price and terra_price != 0:
        return terra_price, terra_url, onelec_url + ' ' + str(onelec_price)
    if onelec_price != 0:
        return onelec_price, onelec_url, terra_url + ' ' + str(terra_price)
    return terra_price, terra_url, onelec_url + ' ' + str(onelec_price)


def get_PN_from_terra(url: str):
    """
    gets PN from terra using product linf
    :param url: link at product
    :return: Partnumber
    """
    res = requests.get(url)
    soup = BeautifulSoup(res.text)
    pn = soup.find('h1')
    return pn.contents[0].split()[0]

def get_best_price_by_PN(value: str) -> (float, str):
    """
    function gets best price from terra searhing by PN
    :param value: value to search
    :return: price, product id
    """
    url = terra_base + "search?text=" + value
    r = requests.get(url)
    link = r.url
    products = []
    if 'catalog' not in link:
        soup = BeautifulSoup(r.text)
        links = soup.find('ul', {'class': "search-list"})
        link = links.contents[1].contents[1].attrs['href']
        get_product_data(link, products)
    else:
        get_product_data(link.split('ru/')[1], products)
    if products:
        best_price, best_url, _ =  get_min_price_quantity_data(products, 1, 5)
        best_url = terra_base + 'product/' + best_url
        if best_url:
            PN = get_PN_from_terra(best_url)
            price_onelec, url_onelec = get_onelec_pn(PN.lower())
            if price_onelec > 0 and price_onelec < best_price:
                best_price = price_onelec, best_url = url_onelec
        return best_price, best_url
    else:
        return -1, ""


def write_results(results: dict):
    """
    write results to results.xlsx workbook
    :param results: dictionary with results
    :return:
    """
    wb = Workbook()
    ws1 = wb.active
    ws1['A1'] = "Value"
    ws1['B1'] = "Price"
    ws1["C1"] = "URL"
    i = 2
    for value in results.keys():
        ws1['A%i' % i] = value
        ws1['B%i' % i] = results[value][0]
        ws1['C%i' % i] = results[value][1]
        i+=1
    wb.save(filename = 'Results.xlsx')


def main(filename, start=1, end=100):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    results = {}
    for i in range (int(start)+1, int(end)+2):
        value = sheet['A%i' % i].value
        component = sheet['B%i' %i].value
        footprint = sheet['C%i' %i].value
        PN = sheet['D%i' %i].value
        if component == "CAP" or component == 'RES':
            search_links = get_search_links_for_position(component, value, footprint)
            if search_links:
                products = []
                for link in search_links:
                    get_product_data(link, products)
            if products:
                quantity = 1
                best_price, best_price_id, _ = get_min_price_quantity_data(products, quantity, 5)
                results[value] = [best_price, terra_base + best_price_id]
        if PN:
            best_price, best_url, comment = get_best_price_from_onelec_terra_by_pn(PN)
            if best_price != 0:
                results[value] = [best_price, best_url]
            else:
                results[value] = [-1, best_url]
        if component == 'PN':
            best_price, best_price_url = get_best_price_by_PN(value)
            results[value] = [best_price, best_price_url]
        if value not in results.keys():
            results[value] = [-1, ""]
        write_results(results)
        print(value, results[value])
    #write_results(results)




if __name__ == '__main__':
    if len(sys.argv) not in [2, 4]:
        print('Please specify file name (necessary) and range (start and end line, optional)')
    else:
        main(*sys.argv[1:])


